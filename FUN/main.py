import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb= openpyxl.load_workbook("Book1.xlsx")

ws= wb.active

ws['B12']= "=sum(B2:B9)"
ws['b13']= "=average(B2:B9)"

ws['d1']= 'Balance after a year'
ws['d1'].font = Font(bold=True, name='arial',size=11)
ws['E1']= 'Interest Accrued after a year'
ws['E1'].font = Font(bold=True, name='arial',size=11)
ws['F1']= 'Balance after 2 years'
ws['F1'].font = Font(bold=True, name='arial',size=11)




for i in range(2,10):
   Balance=  ws.cell(row=i,column=2).value
   Interest= ws.cell(row=i, column=3).value
   Final_Balance= (Balance*Interest)+Balance
   Interest_Amount= (Balance*Interest)
   Balance2 = (Final_Balance*Interest)+Final_Balance
   ws.cell(row=i,column=4).value = Final_Balance
   ws.cell(row=i, column=5).value= Interest_Amount
   ws.cell(row=i, column=6).value= Balance2

length_of_columnD = len(ws['D1'].value)
ws.column_dimensions['D'].width= length_of_columnD

length_of_columnE = len(ws['E1'].value)
ws.column_dimensions['E'].width= length_of_columnE

length_of_columnF = len(ws['F1'].value)
ws.column_dimensions['F'].width= length_of_columnF

for letter in ['A','B']:
   max_width = 0

   for row_number in range(1, ws.max_row +1):
      if len(str(ws[f'{letter}{row_number}'].value)) > max_width:
         max_width = len(str(ws[f'{letter}{row_number}'].value))

   ws.column_dimensions[letter].width = max_width + 2


# (balance*interest)+balance


wb.save("output.xlsx")